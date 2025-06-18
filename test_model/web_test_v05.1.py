from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from itertools import product
import string
import time
import csv
import logging
import os
import random
import re
import sys
import json
from bs4 import BeautifulSoup

# ตั้งค่า logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
DEBUG_MODE = False

def set_debug_mode(enabled=True):
    """เปิด/ปิด debug mode"""
    global DEBUG_MODE
    DEBUG_MODE = enabled
    logger.setLevel(logging.DEBUG if enabled else logging.INFO)

def debug_print(message):
    """แสดงข้อความ debug"""
    if DEBUG_MODE:
        print(f"🔍 DEBUG: {message}")

def generate_keywords():
    """อ่านรายชื่อยาจากไฟล์เพื่อใช้เป็นคำค้นหา โดยดึงเฉพาะชื่อยาที่อยู่ใน double quotes"""
    try:
        import re
        with open('Name_Ya_all.txt', 'r', encoding='utf-8') as f:
            content = f.read()
            
            # นับจำนวนบรรทัดทั้งหมด
            total_lines = len(content.splitlines())
            logger.info(f"จำนวนบรรทัดทั้งหมดในไฟล์: {total_lines}")
            
            # ดึงเฉพาะข้อความที่อยู่ใน double quotes
            drug_names = re.findall(r'"([^"]*)"', content)
            logger.info(f"จำนวนชื่อยาที่อยู่ในเครื่องหมาย quotes: {len(drug_names)}")
            
            # เก็บชื่อยาที่ถูกตัดทิ้งเพื่อตรวจสอบ
            skipped_names = []
            
            # ทำความสะอาดชื่อยา - เอาเฉพาะชื่อหลักก่อนวงเล็บ
            cleaned_names = []
            for name in drug_names:
                # แยกเอาชื่อหลักก่อนวงเล็บ
                main_name = name.split('(')[0].strip()
                # ถ้าชื่อไม่ว่างเปล่า ให้เพิ่มเข้า list
                if main_name:
                    cleaned_names.append(main_name)
                else:
                    skipped_names.append(name)
            
            logger.info(f"จำนวนชื่อยาหลังทำความสะอาด: {len(cleaned_names)}")
            
            if skipped_names:
                logger.warning(f"มีชื่อยาที่ถูกข้าม {len(skipped_names)} รายการ")
                # บันทึกชื่อยาที่ถูกข้ามลงไฟล์
                with open('skipped_drug_names.txt', 'w', encoding='utf-8') as f:
                    for name in skipped_names:
                        f.write(f"{name}\n")
                logger.info("บันทึกรายชื่อยาที่ถูกข้ามไว้ใน skipped_drug_names.txt")
            
            # ตรวจสอบและบันทึกข้อมูลชื่อยาที่ซ้ำกัน แต่ยังคงเก็บไว้ทั้งหมด
            name_counts = {}
            for name in cleaned_names:
                name_counts[name] = name_counts.get(name, 0) + 1
            
            duplicates = {name: count for name, count in name_counts.items() if count > 1}
            if duplicates:
                logger.info(f"พบชื่อยาที่ซ้ำกัน {len(duplicates)} รายการ (แต่จะเก็บไว้ทั้งหมด)")
                # บันทึกชื่อยาที่ซ้ำกันลงไฟล์
                with open('duplicate_drug_names.txt', 'w', encoding='utf-8') as f:
                    for name, count in sorted(duplicates.items()):
                        f.write(f"{name} (ซ้ำ {count} ครั้ง)\n")
                logger.info("บันทึกรายชื่อยาที่ซ้ำกันไว้ใน duplicate_drug_names.txt")
            
            return cleaned_names  # ส่งคืนชื่อยาทั้งหมด รวมทั้งชื่อที่ซ้ำกัน
            
    except Exception as e:
        logger.error(f"เกิดข้อผิดพลาดในการอ่านไฟล์รายชื่อยา: {e}")
        # ถ้าเกิดข้อผิดพลาด ให้ใช้วิธีเดิมเป็น fallback
        return [''.join(p) for p in product(string.ascii_lowercase, repeat=2)]

def print_drug_data(data, index=None):
    """แสดงข้อมูลยาในรูปแบบที่อ่านง่าย"""
    print("\n" + "="*80)
    print(f"💊 ข้อมูลยาที่ {index}" if index else "💊 ข้อมูลยา")
    print("="*80)
    
    # ข้อมูลพื้นฐาน
    basic_info = ["ชื่อสามัญ", "ชื่อการค้า", "รูปแบบยา"]
    for key in basic_info:
        value = data.get(key, "")
        status = "🔹" if value else "❌"
        print(f"{status} {key}: {value if value else '(ไม่มีข้อมูล)'}")
    
    print("-" * 80)
    
    # ข้อมูลรายละเอียด
    detail_keys = [
        "ยานี้ใช้สำหรับ", "วิธีการใช้ยา", "สิ่งที่ควรแจ้งให้แพทย์หรือเภสัชกรทราบ",
        "ทำอย่างไรหากลืมรับประทานยาหรือใช้ยา", "อาการไม่พึงประสงค์ทั่วไป",
        "อาการไม่พึงประสงค์ที่ต้องแจ้งแพทย์หรือเภสัชกรทันที", "การเก็บรักษายา"
    ]
    
    for key in detail_keys:
        value = data.get(key, "")
        print(f"\n🔸 {key}:")
        if value:
            for line in value.split('\n'):
                if line.strip():
                    print(f"   {line.strip()}")
        else:
            print("   (ไม่มีข้อมูล)")
    
    if data.get("URL"):
        print(f"\n🔗 URL: {data['URL']}")
    print("=" * 80)

def extract_full_text_with_linebreaks(driver):
    """ดึงข้อมูลจาก HTML พร้อมแทนที่ <br> ด้วย \n"""
    soup = BeautifulSoup(driver.page_source, "html.parser")
    for br in soup.find_all("br"):
        br.replace_with("\n")
    return soup.get_text(separator="\n")

def print_summary_stats(all_drugs, keyword, current_index):
    """แสดงสถิติสรุป"""
    complete_data = sum(1 for drug in all_drugs if any(drug.get(field, "").strip() 
                       for field in ["ชื่อสามัญ", "ชื่อการค้า", "ยานี้ใช้สำหรับ"]))
    
    print(f"\n📊 สรุปข้อมูล (Keyword: {keyword})")
    print(f"📈 รวมยาที่ดึงมาได้: {len(all_drugs)} รายการ")
    print(f"✅ ข้อมูลครบถ้วน: {complete_data} รายการ")
    print(f"⚠️ ข้อมูลไม่ครบ: {len(all_drugs) - complete_data} รายการ")
    print("-" * 50)

def extract_drug_usage_patterns(text):
    """ดึงข้อมูลการใช้ยาด้วยรูปแบบที่หลากหลาย"""
    patterns = {
        'general': [
            r'(?:ยานี้ใช้|ใช้สำหรับ|ใช้เพื่อ|ใช้รักษา)[^.]*?(?=\.|$)',
            r'(?:มีข้อบ่งใช้|มีสรรพคุณ)[^.]*?(?=\.|$)',
            r'(?:ใช้บรรเทา|ใช้บำบัด)[^.]*?(?=\.|$)'
        ],
        'specific_conditions': [
            r'(?:รักษาโรค|บรรเทาอาการ)[^.]*?(?=\.|$)',
            r'(?:ควบคุมอาการ|ลดอาการ)[^.]*?(?=\.|$)',
            r'(?:ป้องกัน|รักษา)[^.]*?(?=\.|$)'
        ],
        'combinations': [
            r'(?:ใช้ร่วมกับ|ใช้เสริมกับ)[^.]*?(?=\.|$)',
            r'(?:อาจใช้ร่วม|สามารถใช้ร่วม)[^.]*?(?=\.|$)'
        ],
        'special_cases': [
            r'(?:กรณีพิเศษ|ในกรณี)[^.]*?(?=\.|$)',
            r'(?:ข้อควรระวัง|ข้อห้ามใช้)[^.]*?(?=\.|$)'
        ]
    }
    
    results = []
    seen = set()
    
    for category, pattern_list in patterns.items():
        for pattern in pattern_list:
            matches = re.finditer(pattern, text, re.IGNORECASE | re.MULTILINE)
            for match in matches:
                content = match.group().strip()
                if len(content) > 20 and content not in seen:
                    results.append(content)
                    seen.add(content)
    
    return results

def extract_administration_patterns(text):
    """ดึงข้อมูลวิธีการใช้ยาด้วยรูปแบบที่หลากหลาย"""
    patterns = {
        'oral': [
            r'(?:รับประทาน|กิน)[^.]*?(?:ครั้งละ|วันละ|ทุก|มื้อ)[^.]*?(?=\.|$)',
            r'(?:ควรรับประทาน|ให้รับประทาน)[^.]*?(?=\.|$)'
        ],
        'injection': [
            r'(?:ฉีด|ให้ยาทาง)[^.]*?(?:เข้า|ทาง|ใต้)[^.]*?(?=\.|$)',
            r'(?:ขนาดยาฉีด|วิธีฉีด)[^.]*?(?=\.|$)'
        ],
        'topical': [
            r'(?:ทา|ป้าย|พ่น)[^.]*?(?:บริเวณ|ผิวหนัง|แผล)[^.]*?(?=\.|$)',
            r'(?:วิธีทา|การทา)[^.]*?(?=\.|$)'
        ],
        'timing': [
            r'(?:ทุก|ก่อน|หลัง)[^.]*?(?:มื้ออาหาร|นาที|ชั่วโมง)[^.]*?(?=\.|$)',
            r'(?:เวลา|ช่วงเวลา)[^.]*?(?=\.|$)'
        ],
        'dosage': [
            r'(?:ขนาดยา|ปริมาณ)[^.]*?(?:ตาม|ขึ้นกับ)[^.]*?(?=\.|$)',
            r'(?:ผู้ใหญ่|เด็ก)[^.]*?(?:รับประทาน|ใช้)[^.]*?(?=\.|$)'
        ],
        'special_instructions': [
            r'(?:ข้อควรระวัง|คำแนะนำ)[^.]*?(?=\.|$)',
            r'(?:หมายเหตุ|ข้อสังเกต)[^.]*?(?=\.|$)'
        ]
    }
    
    results = []
    seen = set()
    
    for category, pattern_list in patterns.items():
        for pattern in pattern_list:
            matches = re.finditer(pattern, text, re.IGNORECASE | re.MULTILINE)
            for match in matches:
                content = match.group().strip()
                if len(content) > 20 and content not in seen:
                    results.append(content)
                    seen.add(content)
    
    return results

def clean_and_format_content(content_list):
    """ทำความสะอาดและจัดรูปแบบเนื้อหา"""
    cleaned = []
    seen = set()
    
    for item in content_list:
        # ทำความสะอาดข้อความ
        clean_item = re.sub(r'\s+', ' ', item).strip()
        clean_item = re.sub(r'^[•\-\*\s]+', '', clean_item)
        
        # ตัดคำที่ไม่จำเป็น
        unwanted = ["ได้แก่", "เช่น", "อาทิ", "คือ", "ดังนี้"]
        for word in unwanted:
            clean_item = clean_item.replace(word, "")
        
        # ตรวจสอบความยาวและความซ้ำซ้อน
        if len(clean_item) > 20 and clean_item not in seen:
            cleaned.append(clean_item)
            seen.add(clean_item)
    
    return cleaned

def extract_drug_usage_section(lines, start_idx, section_indices):
    """ดึงข้อมูล 'ยานี้ใช้สำหรับ' แบบครอบคลุม"""
    try:
        # หาขอบเขตของส่วน
        end_idx = find_section_end(lines, start_idx, section_indices)
        
        # รวบรวมเนื้อหา
        content = gather_section_content(lines, start_idx, end_idx)
        
        if not content:
            return ""
        
        # ประมวลผลเนื้อหา
        combined_text = " ".join(content)
        
        # ดึงข้อมูลด้วยรูปแบบต่างๆ
        usage_patterns = extract_drug_usage_patterns(combined_text)
        
        # ทำความสะอาดและจัดรูปแบบ
        cleaned_patterns = clean_and_format_content(usage_patterns)
        
        if cleaned_patterns:
            return "\n".join([f"* {pattern}" for pattern in cleaned_patterns])
        
        return ""
        
    except Exception as e:
        debug_print(f"Usage extraction error: {e}")
        return ""

def extract_administration_section(lines, start_idx, section_indices):
    """ดึงข้อมูล 'วิธีการใช้ยา' แบบครอบคลุม"""
    try:
        # หาขอบเขตของส่วน
        end_idx = find_section_end(lines, start_idx, section_indices)
        
        # รวบรวมเนื้อหา
        content = gather_section_content(lines, start_idx, end_idx)
        
        if not content:
            return ""
        
        # ประมวลผลเนื้อหา
        combined_text = " ".join(content)
        debug_print(f"Administration content length: {len(combined_text)}")
        
        # ดึงข้อมูลด้วยรูปแบบต่างๆ
        admin_patterns = extract_administration_patterns(combined_text)
        
        # ทำความสะอาดและจัดรูปแบบ
        cleaned_patterns = clean_and_format_content(admin_patterns)
        
        if cleaned_patterns:
            return "\n".join([f"* {pattern}" for pattern in cleaned_patterns])
        
        # ถ้าไม่พบรูปแบบใหม่ ใช้วิธีเดิม
        bullets = extract_administration_bullets_comprehensive(combined_text)
        if bullets:
            return "\n".join([f"* {bullet}" for bullet in bullets])
        
        return ""
        
    except Exception as e:
        debug_print(f"Administration extraction error: {e}")
        return ""

def find_section_end(lines, start_idx, section_indices):
    """หาจุดสิ้นสุดของส่วน"""
    end_idx = len(lines)
    next_section_keywords = [
        "วิธีการใช้ยา", "สิ่งที่ควรแจ้ง", "ทำอย่างไร",
        "อาการไม่พึงประสงค์", "การเก็บรักษายา"
    ]
    
    # หาจุดสิ้นสุดจาก section_indices
    for section_name, idx in section_indices.items():
        if idx > start_idx:
            end_idx = min(end_idx, idx)
    
    # หาจุดสิ้นสุดจากคำสำคัญ
    for i in range(start_idx + 1, end_idx):
        if i >= len(lines):
            break
        if any(keyword in lines[i] for keyword in next_section_keywords):
            end_idx = i
            break
    
    return end_idx

def gather_section_content(lines, start_idx, end_idx):
    """รวบรวมเนื้อหาของส่วน"""
    content = []
    current_idx = start_idx + 1
    
    while current_idx < min(end_idx, len(lines)):
        line = lines[current_idx].strip()
        
        # ข้ามบรรทัดว่าง
        if not line:
            current_idx += 1
            continue
        
        # ตรวจสอบ unwanted line
        if not is_unwanted_line(line):
            # จัดการกับ HTML list
            if line.startswith('<li>'):
                list_item = []
                while current_idx < len(lines) and '</li>' not in lines[current_idx]:
                    list_item.append(lines[current_idx])
                    current_idx += 1
                if current_idx < len(lines):
                    list_item.append(lines[current_idx])
                content.append('\n'.join(list_item))
            else:
                content.append(line)
        
        current_idx += 1
    
    return content

def extract_trade_name_comprehensive(driver, soup):
    """ดึงชื่อการค้าแบบครอบคลุม"""
    trade_name = ""
    
    try:
        # 1. ดึงจาก h1 หรือ h2 ที่มีคำว่า "ชื่อการค้า"
        for tag in ["h1", "h2"]:
            headers = soup.find_all(tag)
            for header in headers:
                if "ชื่อการค้า" in header.text:
                    # ดึงข้อความหลังคำว่า "ชื่อการค้า"
                    text = header.text.split("ชื่อการค้า")[-1].strip()
                    if text and len(text) > 1:
                        return clean_single_line_content(text)
        
        # 2. ดึงจาก div.bs-callout ที่มีคำว่า "ชื่อการค้า"
        callouts = soup.find_all("div", class_="bs-callout")
        for callout in callouts:
            if "ชื่อการค้า" in callout.text:
                # ดึงข้อความหลังคำว่า "ชื่อการค้า"
                text = callout.text.split("ชื่อการค้า")[-1].strip()
                if text and len(text) > 1:
                    return clean_single_line_content(text)
        
        # 3. ดึงจาก strong tag ที่มีคำว่า "ชื่อการค้า"
        strongs = soup.find_all("strong")
        for strong in strongs:
            if "ชื่อการค้า" in strong.text:
                # หาข้อความถัดไป
                next_sibling = strong.next_sibling
                if next_sibling:
                    text = next_sibling.strip()
                    if text and len(text) > 1:
                        return clean_single_line_content(text)
        
        # 4. ดึงจาก table ที่มีคำว่า "ชื่อการค้า"
        tables = soup.find_all("table")
        for table in tables:
            rows = table.find_all("tr")
            for row in rows:
                cells = row.find_all("td")
                if len(cells) >= 2 and "ชื่อการค้า" in cells[0].text:
                    text = cells[1].text.strip()
                    if text and len(text) > 1:
                        return clean_single_line_content(text)
        
        # 5. ดึงจากหัวข้อเพจ
        title = driver.title
        if title:
            # ตัดคำว่า "ยา" ออกถ้ามี
            title = title.replace("ยา", "").strip()
            # ตัดข้อความหลัง | หรือ - ถ้ามี
            for separator in ["|", "-", ":", "/"]:
                if separator in title:
                    title = title.split(separator)[0].strip()
            if title and len(title) > 1:
                return clean_single_line_content(title)
        
        # 6. ดึงจาก meta tags
        meta_title = soup.find("meta", property="og:title")
        if meta_title and meta_title.get("content"):
            text = meta_title["content"].strip()
            # ตัดคำว่า "ยา" ออกถ้ามี
            text = text.replace("ยา", "").strip()
            # ตัดข้อความหลัง | หรือ - ถ้ามี
            for separator in ["|", "-", ":", "/"]:
                if separator in text:
                    text = text.split(separator)[0].strip()
            if text and len(text) > 1:
                return clean_single_line_content(text)
                
    except Exception as e:
        if DEBUG_MODE:
            debug_print(f"Trade name extraction error: {e}")
    
    return trade_name

def extract_drug_details_fixed(driver):
    """ฟังก์ชันดึงข้อมูลยาหลัก"""
    data = {
        "ชื่อสามัญ": "", "ชื่อการค้า": "", "รูปแบบยา": "", "ยานี้ใช้สำหรับ": "",
        "วิธีการใช้ยา": "", "สิ่งที่ควรแจ้งให้แพทย์หรือเภสัชกรทราบ": "",
        "ทำอย่างไรหากลืมรับประทานยาหรือใช้ยา": "", "อาการไม่พึงประสงค์ทั่วไป": "",
        "อาการไม่พึงประสงค์ที่ต้องแจ้งแพทย์หรือเภสัชกรทันที": "", "การเก็บรักษายา": "",
        "URL": driver.current_url
    }

    try:
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        time.sleep(2)

        # ใช้ BeautifulSoup ในการแยกวิเคราะห์ HTML
        html_content = driver.page_source
        
        # ลบ <b>มีดังนี้</b> และรูปแบบที่คล้ายกันออก
        html_content = re.sub(r'<b>\s*มีดังนี้\s*</b>', '', html_content)
        html_content = re.sub(r'<strong>\s*มีดังนี้\s*</strong>', '', html_content)
        html_content = re.sub(r'<b>\s*ดังนี้\s*</b>', '', html_content)
        html_content = re.sub(r'<strong>\s*ดังนี้\s*</strong>', '', html_content)
        
        soup = BeautifulSoup(html_content, "html.parser")

        # ดึงชื่อสามัญก่อน - เพิ่มวิธีการดึงใหม่
        data["ชื่อสามัญ"] = extract_generic_name_comprehensive(driver, soup)
        
        # ดึงชื่อการค้า
        data["ชื่อการค้า"] = extract_trade_name_comprehensive(driver, soup)
        
        # ดึงข้อมูลจาก div.bs-callout
        callouts = soup.find_all("div", class_="bs-callout bs-callout-warning")
        
        for callout in callouts:
            # ดึงหัวข้อจาก h4
            header = callout.find("h4")
            if not header:
                continue
                
            # ทำความสะอาดหัวข้อ
            header_text = header.get_text(strip=True)
            header_text = re.sub(r'\s*<[^>]+>\s*', '', header_text)  # ลบ HTML tags
            header_text = re.sub(r'\s*\([^)]*\)\s*', '', header_text)  # ลบวงเล็บและข้อความในวงเล็บ
            
            # ข้ามส่วน footer และ disclaimer
            if any(skip in header_text.lower() for skip in ["ลิขสิทธิ์", "copyright", "disclaimer"]):
                continue
                
            # หาส่วนที่ตรงกับหัวข้อในข้อมูล
            matching_key = None
            for key in data.keys():
                if key != "URL" and key in header_text:
                    matching_key = key
                    break
                    
            if not matching_key:
                continue
                
            # ดึงเนื้อหา
            content = []
            
            # ดึงจาก bullet points ถ้ามี
            bullets = callout.find_all("li")
            if bullets:
                for bullet in bullets:
                    text = bullet.get_text(strip=True)
                    if text and not any(skip in text.lower() for skip in ["หน้าหลัก", "ติดต่อเรา"]):
                        content.append(text)
            else:
                # ถ้าไม่มี bullet points ดึงข้อความทั้งหมด
                # ดึงข้อความที่อยู่หลัง h4 โดยตรง
                next_element = header.next_sibling
                if next_element:
                    text = next_element.strip()
                    if text:
                        content.append(text)
            
            # จัดรูปแบบและบันทึกข้อมูล
            if content:
                if matching_key in ["ชื่อสามัญ", "ชื่อการค้า", "รูปแบบยา"]:
                    data[matching_key] = clean_single_line_content(" ".join(content))
                else:
                    data[matching_key] = format_as_bullets(content)
        
        # ถ้ายังไม่ได้ข้อมูลบางส่วน ลองใช้วิธีอื่น
        if not any(data[key] for key in ["ชื่อสามัญ", "ชื่อการค้า", "รูปแบบยา"]):
            extract_from_html_elements(driver, data)
            
        # ทำความสะอาดข้อมูลสุดท้าย
        for key, value in data.items():
            if isinstance(value, str) and value:
                cleaned = value.strip()
                if cleaned in ["", "-", "*", "**"] or len(cleaned) <= 3:
                    data[key] = ""
                else:
                    data[key] = cleaned

    except Exception as e:
        if DEBUG_MODE:
            debug_print(f"Error in extract_drug_details_fixed: {e}")

    return data

def extract_generic_name_comprehensive(driver, soup):
    """ดึงชื่อสามัญแบบครอบคลุม"""
    try:
        generic_name = ""
        
        # 1. ดึงจาก div.bs-callout ที่มีคำว่า "ชื่อสามัญ"
        callouts = soup.find_all("div", class_="bs-callout")
        for callout in callouts:
            if "ชื่อสามัญ" in callout.text:
                # ดึงข้อความหลังคำว่า "ชื่อสามัญ"
                text = callout.text.split("ชื่อสามัญ")[-1].strip()
                if text and len(text) > 1:
                    return clean_single_line_content(text)
        
        # 2. ดึงจาก h1, h2, h3, h4 ที่มีคำว่า "ชื่อสามัญ"
        for tag in ["h1", "h2", "h3", "h4"]:
            headers = soup.find_all(tag)
            for header in headers:
                if "ชื่อสามัญ" in header.text:
                    # หาข้อความถัดไป
                    next_sibling = header.next_sibling
                    if next_sibling:
                        text = next_sibling.strip()
                        if text and len(text) > 1:
                            return clean_single_line_content(text)
        
        # 3. ดึงจาก strong tag ที่มีคำว่า "ชื่อสามัญ"
        strongs = soup.find_all("strong")
        for strong in strongs:
            if "ชื่อสามัญ" in strong.text:
                # หาข้อความถัดไป
                next_sibling = strong.next_sibling
                if next_sibling:
                    text = next_sibling.strip()
                    if text and len(text) > 1:
                        return clean_single_line_content(text)
        
        # 4. ดึงจาก table ที่มีคำว่า "ชื่อสามัญ"
        tables = soup.find_all("table")
        for table in tables:
            rows = table.find_all("tr")
            for row in rows:
                cells = row.find_all("td")
                if len(cells) >= 2 and "ชื่อสามัญ" in cells[0].text:
                    text = cells[1].text.strip()
                    if text and len(text) > 1:
                        return clean_single_line_content(text)
        
        # 5. ดึงจาก meta tags
        meta_generic = soup.find("meta", {"name": "generic-name"}) or soup.find("meta", {"property": "og:generic-name"})
        if meta_generic and meta_generic.get("content"):
            text = meta_generic["content"].strip()
            if text and len(text) > 1:
                return clean_single_line_content(text)
        
        # 6. ดึงจาก div ที่มี id หรือ class เกี่ยวกับชื่อสามัญ
        generic_divs = soup.find_all("div", {"id": re.compile(r"generic.*name", re.I)}) + \
                      soup.find_all("div", {"class": re.compile(r"generic.*name", re.I)})
        for div in generic_divs:
            text = div.text.strip()
            if text and len(text) > 1:
                return clean_single_line_content(text)
        
        # 7. ดึงจาก span ที่มีคำว่า "ชื่อสามัญ"
        spans = soup.find_all("span")
        for span in spans:
            if "ชื่อสามัญ" in span.text:
                # หาข้อความถัดไป
                next_sibling = span.next_sibling
                if next_sibling:
                    text = next_sibling.strip()
                    if text and len(text) > 1:
                        return clean_single_line_content(text)
        
        return generic_name
        
    except Exception as e:
        if DEBUG_MODE:
            debug_print(f"Generic name extraction error: {e}")
        return ""

def extract_structured_data(driver):
    """ดึงข้อมูลจาก structured data"""
    try:
        # ดึงข้อมูลจาก schema.org หรือ JSON-LD
        scripts = driver.find_elements(By.XPATH, '//script[@type="application/ld+json"]')
        structured_data = {}
        
        for script in scripts:
            try:
                json_data = json.loads(script.get_attribute('innerHTML'))
                if isinstance(json_data, dict):
                    if json_data.get("@type") in ["Drug", "Medicine", "MedicalEntity"]:
                        structured_data.update({
                            "ชื่อสามัญ": json_data.get("nonProprietaryName", ""),
                            "ชื่อการค้า": json_data.get("name", ""),
                            "รูปแบบยา": json_data.get("dosageForm", ""),
                            "ยานี้ใช้สำหรับ": json_data.get("indication", "")
                        })
            except:
                continue
                
        return structured_data
    except:
        return {}

def extract_meta_data(driver):
    """ดึงข้อมูลจาก meta tags"""
    try:
        meta_data = {}
        meta_tags = driver.find_elements(By.TAG_NAME, "meta")
        
        for tag in meta_tags:
            try:
                name = tag.get_attribute("name") or tag.get_attribute("property")
                content = tag.get_attribute("content")
                
                if name and content:
                    if "title" in name.lower():
                        meta_data["ชื่อการค้า"] = content
                    elif "description" in name.lower() and "ยานี้ใช้" in content:
                        meta_data["ยานี้ใช้สำหรับ"] = content
            except:
                continue
                
        return meta_data
    except:
        return {}

def merge_structured_data(data, structured_data):
    """ผสานข้อมูลจาก structured data"""
    for key, value in structured_data.items():
        if not data[key] and value:
            data[key] = value

def merge_meta_data(data, meta_data):
    """ผสานข้อมูลจาก meta data"""
    for key, value in meta_data.items():
        if not data[key] and value:
            data[key] = value

def fill_missing_data(driver, data):
    """เติมข้อมูลที่หายไป"""
    try:
        # ถ้าไม่มีชื่อการค้า ลองหาจากหัวข้อหน้า
        if not data["ชื่อการค้า"]:
            try:
                title = driver.title
                if title and ":" in title:
                    data["ชื่อการค้า"] = title.split(":")[0].strip()
            except:
                pass
        
        # ถ้าไม่มีข้อมูลการใช้ยา ลองหาจากเนื้อหาทั้งหมด
        if not data["ยานี้ใช้สำหรับ"]:
            try:
                body_text = driver.find_element(By.TAG_NAME, "body").text
                usage_matches = re.findall(r"(?:ยานี้ใช้|ใช้สำหรับ|ใช้รักษา).*?(?=\n|$)", body_text)
                if usage_matches:
                    data["ยานี้ใช้สำหรับ"] = format_as_bullets(usage_matches)
            except:
                pass
                
        # ถ้าไม่มีรูปแบบยา ลองหาจากข้อความที่เกี่ยวข้อง
        if not data["รูปแบบยา"]:
            try:
                body_text = driver.find_element(By.TAG_NAME, "body").text.lower()
                
                # รูปแบบยาและคำที่เกี่ยวข้อง
                form_patterns = {
                    "เจล": [r"(?:ยา)?เจล", r"gel", r"เจลทา", r"เจลใช้เฉพาะที่", r"(?:ยา)?เจล\s*(?:ทา)?", r"gel\s*(?:form)?", r"topical gel"],
                    "เม็ด": [r"ยาเม็ด", r"tablet", r"เม็ด"],
                    "แคปซูล": [r"แคปซูล", r"capsule"],
                    "น้ำ": [r"ยาน้ำ", r"น้ำเชื่อม", r"น้ำแขวนตะกอน", r"syrup", r"suspension"],
                    "ครีม": [r"ครีม", r"cream"],
                    "ขี้ผึ้ง": [r"ขี้ผึ้ง", r"ointment"],
                    "ยาฉีด": [r"ฉีด", r"injection"],
                    "ยาพ่น": [r"พ่น", r"spray"]
                }
                
                # ตรวจสอบแต่ละรูปแบบยา
                for form, patterns in form_patterns.items():
                    for pattern in patterns:
                        if re.search(pattern, body_text, re.IGNORECASE):
                            data["รูปแบบยา"] = form
                            return
                            
                # ถ้ายังไม่พบ ลองหาจาก class หรือ id ที่เกี่ยวข้อง
                form_elements = driver.find_elements(By.CSS_SELECTOR, "[class*='form'], [class*='type'], [id*='form'], [id*='type']")
                for element in form_elements:
                    element_text = element.text.lower()
                    for form, patterns in form_patterns.items():
                        if any(re.search(pattern, element_text, re.IGNORECASE) for pattern in patterns):
                            data["รูปแบบยา"] = form
                            return
            except:
                pass
    except:
        pass

def extract_from_html_elements(driver, data):
    """ดึงข้อมูลจาก HTML elements โดยตรง"""
    try:
        # 1. ดึงจาก table
        tables = driver.find_elements(By.TAG_NAME, "table")
        for table in tables:
            rows = table.find_elements(By.TAG_NAME, "tr")
            for row in rows:
                cells = row.find_elements(By.TAG_NAME, "td")
                if len(cells) >= 2:
                    header = cells[0].text.strip()
                    content = cells[1].text.strip()
                    
                    for section in data.keys():
                        if section != "URL" and not data[section] and section in header and content:
                            data[section] = clean_single_line_content(content) if section in ["ชื่อสามัญ", "ชื่อการค้า", "รูปแบบยา"] else format_as_bullets([content])
        
        # 2. ดึงจาก div ที่มี class หรือ id เฉพาะ
        target_classes = ["drug-info", "drug-details", "drug-content", "content", "main"]
        for class_name in target_classes:
            elements = driver.find_elements(By.CLASS_NAME, class_name)
            for element in elements:
                extract_from_element_content(element, data)
        
        # 3. ดึงจาก heading elements
        headings = []
        for level in range(1, 7):
            headings.extend(driver.find_elements(By.TAG_NAME, f"h{level}"))
        
        for heading in headings:
            try:
                heading_text = heading.text.strip()
                for section in data.keys():
                    if section != "URL" and not data[section] and section in heading_text:
                        # หาเนื้อหาจาก element ถัดไป
                        next_element = get_next_sibling_element(driver, heading)
                        if next_element:
                            content = next_element.text.strip()
                            if content:
                                data[section] = clean_single_line_content(content) if section in ["ชื่อสามัญ", "ชื่อการค้า", "รูปแบบยา"] else format_as_bullets([content])
            except:
                continue
        
        # 4. ดึงจาก list elements
        lists = driver.find_elements(By.TAG_NAME, "ul") + driver.find_elements(By.TAG_NAME, "ol")
        for list_element in lists:
            try:
                # หา heading ที่อยู่ก่อนหน้า
                prev_element = get_previous_sibling_element(driver, list_element)
                if prev_element:
                    prev_text = prev_element.text.strip()
                    for section in data.keys():
                        if section != "URL" and not data[section] and section in prev_text:
                            list_items = list_element.find_elements(By.TAG_NAME, "li")
                            content = [item.text.strip() for item in list_items if item.text.strip()]
                            if content:
                                data[section] = format_as_bullets(content)
            except:
                continue
        
        # 5. ดึงจาก structured elements
        sections = driver.find_elements(By.TAG_NAME, "section")
        for section_element in sections:
            try:
                section_id = section_element.get_attribute("id") or ""
                section_class = section_element.get_attribute("class") or ""
                
                for section in data.keys():
                    if section != "URL" and not data[section]:
                        section_lower = section.lower()
                        if (section_lower in section_id.lower() or 
                            section_lower in section_class.lower() or 
                            section in section_element.text):
                            content = section_element.text.strip()
                            if content:
                                data[section] = clean_single_line_content(content) if section in ["ชื่อสามัญ", "ชื่อการค้า", "รูปแบบยา"] else format_as_bullets([content])
            except:
                continue
                
    except Exception as e:
        debug_print(f"HTML element extraction failed: {e}")

def extract_from_element_content(element, data):
    """ดึงข้อมูลจากเนื้อหาของ element"""
    try:
        element_text = element.text.strip()
        if not element_text:
            return
            
        # แยกเนื้อหาเป็นบรรทัด
        lines = [line.strip() for line in element_text.split('\n') if line.strip()]
        
        for i, line in enumerate(lines):
            for section in data.keys():
                if section != "URL" and not data[section] and section in line:
                    # หาเนื้อหาในบรรทัดถัดไป
                    content_lines = []
                    j = i + 1
                    while j < len(lines) and not any(key in lines[j] for key in data.keys() if key != "URL"):
                        content_lines.append(lines[j])
                        j += 1
                    
                    if content_lines:
                        content = "\n".join(content_lines)
                        data[section] = clean_single_line_content(content) if section in ["ชื่อสามัญ", "ชื่อการค้า", "รูปแบบยา"] else format_as_bullets(content_lines)
    except:
        pass

def get_next_sibling_element(driver, element):
    """หา element ถัดไป"""
    try:
        return driver.execute_script("""
            function getNextSibling(elem) {
                var sibling = elem.nextElementSibling;
                while (sibling) {
                    if (sibling.offsetParent !== null) {  // check if element is visible
                        return sibling;
                    }
                    sibling = sibling.nextElementSibling;
                }
                return null;
            }
            return getNextSibling(arguments[0]);
        """, element)
    except:
        return None

def get_previous_sibling_element(driver, element):
    """หา element ก่อนหน้า"""
    try:
        return driver.execute_script("""
            function getPreviousSibling(elem) {
                var sibling = elem.previousElementSibling;
                while (sibling) {
                    if (sibling.offsetParent !== null) {  // check if element is visible
                        return sibling;
                    }
                    sibling = sibling.previousElementSibling;
                }
                return null;
            }
            return getPreviousSibling(arguments[0]);
        """, element)
    except:
        return None

def safe_click_element(driver, element, max_retries=3):
    """ปรับปรุงการคลิก element"""
    for retry in range(max_retries):
        try:
            current_url = driver.current_url
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
            time.sleep(random.uniform(0.5, 1.0))
            
            driver.execute_script("arguments[0].click();", element)
            time.sleep(random.uniform(1.5, 2.5))
            
            if driver.current_url != current_url:
                return True
            else:
                element.click()
                time.sleep(random.uniform(1.5, 2.5))
                if driver.current_url != current_url:
                    return True
                    
        except StaleElementReferenceException:
            time.sleep(1)
            return False
        except Exception as e:
            debug_print(f"Click failed, retry {retry + 1}: {e}")
            time.sleep(random.uniform(1, 2))
    
    return False

def safe_navigate_back(driver, max_retries=3):
    """ปรับปรุงการย้อนกลับ"""
    for retry in range(max_retries):
        try:
            current_url = driver.current_url
            driver.back()
            time.sleep(random.uniform(2, 4))
            
            WebDriverWait(driver, 15).until(lambda d: d.execute_script("return document.readyState") == "complete")
            
            if driver.current_url != current_url:
                return True
                
        except Exception as e:
            debug_print(f"Navigate back failed, retry {retry + 1}: {e}")
            time.sleep(random.uniform(1, 2))
    
    # Fallback
    try:
        driver.get("https://www.yaandyou.net/")
        time.sleep(random.uniform(3, 5))
        return True
    except:
        return False

def wait_for_page_load(driver, timeout=15):
    """รอให้หน้าโหลดเสร็จ"""
    try:
        WebDriverWait(driver, timeout).until(lambda d: d.execute_script("return document.readyState") == "complete")
        time.sleep(random.uniform(0.5, 1.5))
        return True
    except TimeoutException:
        return False

def save_to_csv(data, filename):
    """บันทึกข้อมูลลง CSV"""
    fieldnames = [
        "ชื่อสามัญ", "ชื่อการค้า", "รูปแบบยา", "ยานี้ใช้สำหรับ", "วิธีการใช้ยา",
        "สิ่งที่ควรแจ้งให้แพทย์หรือเภสัชกรทราบ", "ทำอย่างไรหากลืมรับประทานยาหรือใช้ยา",
        "อาการไม่พึงประสงค์ทั่วไป", "อาการไม่พึงประสงค์ที่ต้องแจ้งแพทย์หรือเภสัชกรทันที",
        "การเก็บรักษายา", "URL"
    ]

    try:
        os.makedirs(os.path.dirname(filename) if os.path.dirname(filename) else '.', exist_ok=True)
        
        with open(filename, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for d in data:
                row = {k: d.get(k, "").strip().replace('\n\n', '\n').replace('\r', '') for k in fieldnames}
                writer.writerow(row)
        
        logger.info(f"บันทึกไฟล์ {filename} สำเร็จ ({len(data)} รายการ)")
        return True
    except Exception as e:
        logger.error(f"Error saving CSV {filename}: {e}")
        return False
    
def save_to_json(data, filename):
    """บันทึกข้อมูลลง JSON"""
    try:
        os.makedirs(os.path.dirname(filename) if os.path.dirname(filename) else '.', exist_ok=True)
        
        with open(filename, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        logger.info(f"บันทึกไฟล์ JSON {filename} สำเร็จ ({len(data)} รายการ)")
        return True
    except Exception as e:
        logger.error(f"Error saving JSON {filename}: {e}")
        return False

def save_to_txt(data, filename):
    """บันทึกข้อมูลลง TXT แบบอ่านง่าย"""
    try:
        os.makedirs(os.path.dirname(filename) if os.path.dirname(filename) else '.', exist_ok=True)
        
        with open(filename, "w", encoding="utf-8") as f:
            f.write("=" * 100 + "\n")
            f.write(f"📊 รายงานข้อมูลยา - รวม {len(data)} รายการ\n")
            f.write("=" * 100 + "\n\n")
            
            for i, drug in enumerate(data, 1):
                f.write(f"💊 ยาที่ {i}\n")
                f.write("-" * 80 + "\n")
                
                # ข้อมูลพื้นฐาน
                basic_info = ["ชื่อสามัญ", "ชื่อการค้า", "รูปแบบยา"]
                for key in basic_info:
                    value = drug.get(key, "")
                    status = "🔹" if value else "❌"
                    f.write(f"{status} {key}: {value if value else '(ไม่มีข้อมูล)'}\n")
                
                f.write("-" * 80 + "\n")
                
                # ข้อมูลรายละเอียด
                detail_keys = [
                    "ยานี้ใช้สำหรับ", "วิธีการใช้ยา", "สิ่งที่ควรแจ้งให้แพทย์หรือเภสัชกรทราบ",
                    "ทำอย่างไรหากลืมรับประทานยาหรือใช้ยา", "อาการไม่พึงประสงค์ทั่วไป",
                    "อาการไม่พึงประสงค์ที่ต้องแจ้งแพทย์หรือเภสัชกรทันที", "การเก็บรักษายา"
                ]
                
                for key in detail_keys:
                    value = drug.get(key, "")
                    f.write(f"\n🔸 {key}:\n")
                    if value:
                        for line in value.split('\n'):
                            if line.strip():
                                f.write(f"   {line.strip()}\n")
                    else:
                        f.write("   (ไม่มีข้อมูล)\n")
                
                if drug.get("URL"):
                    f.write(f"\n🔗 URL: {drug['URL']}\n")
                f.write("=" * 100 + "\n\n")
        
        logger.info(f"บันทึกไฟล์ TXT {filename} สำเร็จ ({len(data)} รายการ)")
        return True
    except Exception as e:
        logger.error(f"Error saving TXT {filename}: {e}")
        return False

def save_to_excel(data, filename):
    """บันทึกข้อมูลลง Excel (ต้องติดตั้ง openpyxl ก่อน)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        import pandas as pd
        
        # สร้าง DataFrame
        df = pd.DataFrame(data)
        
        # เรียงลำดับคอลัมน์
        column_order = [
            "ชื่อสามัญ", "ชื่อการค้า", "รูปแบบยา", "ยานี้ใช้สำหรับ", "วิธีการใช้ยา",
            "สิ่งที่ควรแจ้งให้แพทย์หรือเภสัชกรทราบ", "ทำอย่างไรหากลืมรับประทานยาหรือใช้ยา",
            "อาการไม่พึงประสงค์ทั่วไป", "อาการไม่พึงประสงค์ที่ต้องแจ้งแพทย์หรือเภสัชกรทันที",
            "การเก็บรักษายา", "URL"
        ]
        df = df.reindex(columns=column_order)
        
        os.makedirs(os.path.dirname(filename) if os.path.dirname(filename) else '.', exist_ok=True)
        
        # สร้าง workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Drug Information"
        
        # เพิ่มข้อมูลลง worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # จัดรูปแบบหัวตาราง
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # ปรับความกว้างคอลัมน์
        column_widths = {
            'A': 20, 'B': 25, 'C': 15, 'D': 40, 'E': 30,
            'F': 35, 'G': 30, 'H': 25, 'I': 35, 'J': 20, 'K': 50
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Wrap text สำหรับคอลัมน์ที่มีข้อความยาว
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        wb.save(filename)
        logger.info(f"บันทึกไฟล์ Excel {filename} สำเร็จ ({len(data)} รายการ)")
        return True
        
    except ImportError:
        logger.warning("ไม่สามารถบันทึก Excel ได้ กรุณาติดตั้ง: pip install openpyxl pandas")
        return False
    except Exception as e:
        logger.error(f"Error saving Excel {filename}: {e}")
        return False

def save_all_formats(data, base_filename):
    """บันทึกข้อมูลทุกรูปแบบ"""
    base_name = base_filename.replace('.csv', '')
    
    results = {
        'csv': save_to_csv(data, f"{base_name}.csv"),
        'json': save_to_json(data, f"{base_name}.json"),
        'txt': save_to_txt(data, f"{base_name}.txt"),
        'excel': save_to_excel(data, f"{base_name}.xlsx")
    }
    
    success_count = sum(results.values())
    logger.info(f"📁 บันทึกสำเร็จ {success_count}/4 รูปแบบ")
    
    return results

def load_existing_data(filename):
    """โหลดข้อมูลที่มีอยู่แล้ว"""
    if not os.path.exists(filename):
        return []
    
    try:
        with open(filename, "r", encoding="utf-8-sig") as f:
            data = list(csv.DictReader(f))
            logger.info(f"โหลดข้อมูลเดิม: {len(data)} รายการจาก {filename}")
            return data
    except Exception as e:
        logger.warning(f"ไม่สามารถโหลดข้อมูลเดิม: {e}")
        return []

def is_duplicate_drug(new_drug, existing_drugs):
    """ตรวจสอบว่ายาซ้ำหรือไม่"""
    trade_name = new_drug.get("ชื่อการค้า", "").strip()
    generic_name = new_drug.get("ชื่อสามัญ", "").strip()
    url = new_drug.get("URL", "").strip()
    
    if not trade_name and not generic_name:
        return True
    
    def clean_name(name):
        """ทำความสะอาดชื่อยาเพื่อเปรียบเทียบ"""
        if not name:
            return ""
        # ลบวงเล็บและข้อความในวงเล็บ
        name = re.sub(r'\([^)]*\)', '', name)
        # ลบเครื่องหมายและช่องว่างพิเศษ
        name = re.sub(r'[^\w\s]', '', name)
        # แปลงเป็นตัวพิมพ์เล็กและตัดช่องว่าง
        return ' '.join(name.lower().split())
    
    clean_trade = clean_name(trade_name)
    clean_generic = clean_name(generic_name)
    
    for existing in existing_drugs:
        existing_trade = clean_name(existing.get("ชื่อการค้า", ""))
        existing_generic = clean_name(existing.get("ชื่อสามัญ", ""))
        existing_url = existing.get("URL", "").strip()
        
        # ตรวจสอบ URL ถ้ามี
        if url and existing_url and url == existing_url:
            return True
            
        # ตรวจสอบชื่อการค้า
        if clean_trade and existing_trade:
            if clean_trade == existing_trade:
                return True
            # ตรวจสอบกรณีที่อาจมีการสะกดต่างกันเล็กน้อย
            if len(clean_trade) > 3 and (clean_trade in existing_trade or existing_trade in clean_trade):
                return True
        
        # ตรวจสอบชื่อสามัญ
        if clean_generic and existing_generic:
            if clean_generic == existing_generic:
                return True
            # ตรวจสอบกรณีที่อาจมีการสะกดต่างกันเล็กน้อย
            if len(clean_generic) > 3 and (clean_generic in existing_generic or existing_generic in clean_generic):
                return True
    
    return False

def setup_driver():
    """ตั้งค่า Chrome driver"""
    try:
        options = webdriver.ChromeOptions()
        options.add_experimental_option("detach", True)
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("--disable-images")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")
        
        prefs = {
            "profile.managed_default_content_settings.images": 2,
            "profile.default_content_setting_values.notifications": 2,
            "profile.default_content_settings.popups": 0
        }
        options.add_experimental_option("prefs", prefs)
        
        try:
            service = Service(ChromeDriverManager().install())
        except Exception as e:
            logger.error(f"Error installing ChromeDriver: {e}")
            # Try using the default Chrome driver path as fallback
            service = Service("chromedriver")
            
        driver = webdriver.Chrome(service=service, options=options)
        return driver
    except Exception as e:
        logger.error(f"Error setting up Chrome driver: {e}")
        raise

def load_processed_keywords():
    """โหลดรายการ keywords ที่ประมวลผลไปแล้ว"""
    try:
        if os.path.exists("processed_keywords.txt"):
            with open("processed_keywords.txt", "r", encoding="utf-8") as f:
                return set(line.strip() for line in f if line.strip())
        return set()
    except Exception as e:
        logger.warning(f"ไม่สามารถโหลด processed keywords: {e}")
        return set()

def save_processed_keywords(keywords):
    """บันทึกรายการ keywords ที่ประมวลผลไปแล้ว"""
    try:
        with open("processed_keywords.txt", "w", encoding="utf-8") as f:
            for keyword in sorted(keywords):
                f.write(f"{keyword}\n")
        logger.info(f"บันทึก processed keywords: {len(keywords)} รายการ")
    except Exception as e:
        logger.error(f"ไม่สามารถบันทึก processed keywords: {e}")

def main():
    """ฟังก์ชันหลัก"""
    driver = None
    processed_keywords = load_processed_keywords()
    
    try:
        driver = setup_driver()
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        driver.set_page_load_timeout(30)
        
        logger.info("🚀 เริ่มต้นการทำงาน...")
        driver.get("https://www.yaandyou.net/")
        wait_for_page_load(driver)

        # ข้าม SSL warning
        try:
            driver.find_element(By.ID, "details-button").click()
            time.sleep(0.5)
            driver.find_element(By.ID, "proceed-link").click()
            time.sleep(2)
        except NoSuchElementException:
            pass

        all_drugs = load_existing_data("drug_backup.csv")
        keywords = generate_keywords()
        
        # กรองเฉพาะ keywords ที่ยังไม่ได้ประมวลผล
        remaining_keywords = [k for k in keywords if k not in processed_keywords]
        
        logger.info(f"📊 จะประมวลผล {len(remaining_keywords)} keywords ที่เหลือ")
        logger.info(f"📚 มีข้อมูลเดิม {len(all_drugs)} รายการ")
        logger.info(f"✅ ประมวลผลไปแล้ว {len(processed_keywords)} keywords")

        for idx, keyword in enumerate(remaining_keywords):
            if keyword in processed_keywords:
                continue
                
            logger.info(f"🔍 ค้นหา: {keyword} ({idx+1}/{len(remaining_keywords)}) - รวม: {len(all_drugs)} รายการ")
            
            try:
                # ตรวจสอบหน้าหลัก
                if "yaandyou.net" not in driver.current_url or "index.php" not in driver.current_url:
                    driver.get("https://www.yaandyou.net/")
                    wait_for_page_load(driver)
                
                # ค้นหา keyword
                input_box = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.ID, "keyword")))
                input_box.clear()
                input_box.send_keys(keyword)
                time.sleep(random.uniform(1.5, 2.5))

                # รอ autocomplete
                try:
                    WebDriverWait(driver, 8).until(EC.presence_of_element_located((By.XPATH, '//*[@id="results"]/div')))
                    suggestions = driver.find_elements(By.XPATH, '//*[@id="results"]/div')
                except TimeoutException:
                    logger.info(f"ไม่มี autocomplete สำหรับ {keyword}")
                    processed_keywords.add(keyword)
                    continue

                logger.info(f"พบ autocomplete: {len(suggestions)} รายการ")

                # ประมวลผลแต่ละ suggestion
                for i in range(len(suggestions)):
                    try:
                        # รีเฟรช suggestions
                        input_box = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.ID, "keyword")))
                        input_box.clear()
                        input_box.send_keys(keyword)
                        time.sleep(random.uniform(1.5, 2.5))

                        try:
                            WebDriverWait(driver, 8).until(EC.presence_of_element_located((By.XPATH, '//*[@id="results"]/div')))
                            suggestions = driver.find_elements(By.XPATH, '//*[@id="results"]/div')
                        except TimeoutException:
                            break
                            
                        if i >= len(suggestions):
                            continue

                        if not safe_click_element(driver, suggestions[i]):
                            continue
                            
                        wait_for_page_load(driver)

                        if "index_list.php" not in driver.current_url:
                            safe_navigate_back(driver)
                            continue

                        # หาตารางยา
                        rows = []
                        for selector in ['#tableId3 tbody tr', 'table tbody tr', '.table tbody tr', 'tbody tr']:
                            try:
                                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, selector)))
                                all_rows = driver.find_elements(By.CSS_SELECTOR, selector)
                                rows = [row for row in all_rows if row.find_elements(By.TAG_NAME, "td")]
                                if rows:
                                    break
                            except TimeoutException:
                                continue

                        if not rows:
                            safe_navigate_back(driver)
                            continue

                        logger.info(f" → รายการยา: {len(rows)}")

                        # ประมวลผลแต่ละแถวยา
                        for j in range(len(rows)):
                            try:
                                # รีเฟรช rows
                                fresh_rows = []
                                for selector in ['#tableId3 tbody tr', 'table tbody tr', '.table tbody tr', 'tbody tr']:
                                    try:
                                        all_rows = driver.find_elements(By.CSS_SELECTOR, selector)
                                        fresh_rows = [row for row in all_rows if row.find_elements(By.TAG_NAME, "td")]
                                        if fresh_rows:
                                            break
                                    except:
                                        continue

                                if j >= len(fresh_rows):
                                    break

                                # ดึงชื่อการค้า
                                row_cells = fresh_rows[j].find_elements(By.TAG_NAME, "td")
                                trade_name = ""
                                for cell in row_cells:
                                    cell_text = cell.text.strip()
                                    if cell_text and len(cell_text) > 1:
                                        trade_name = cell_text
                                        break

                                if not trade_name or is_duplicate_drug({"ชื่อการค้า": trade_name}, all_drugs):
                                    continue

                                if not safe_click_element(driver, fresh_rows[j]):
                                    continue
                                    
                                wait_for_page_load(driver)

                                # ดึงข้อมูลยา
                                data = extract_drug_details_fixed(driver)
                                
                                if not data["ชื่อการค้า"]:
                                    data["ชื่อการค้า"] = trade_name
                                
                                # ตรวจสอบคุณภาพข้อมูล
                                if ((data["ชื่อสามัญ"] or data["ชื่อการค้า"]) and 
                                    (data["ยานี้ใช้สำหรับ"] or data["วิธีการใช้ยา"])):
                                    
                                    all_drugs.append(data)
                                    print_drug_data(data, len(all_drugs))
                                    logger.info(f"  ✔️ {trade_name} (รวม: {len(all_drugs)})")

                                # กลับหน้า index_list.php
                                safe_navigate_back(driver)
                                wait_for_page_load(driver)

                                time.sleep(random.uniform(0.5, 1.5))

                            except Exception as e:
                                logger.error(f"Error processing drug row {j}: {e}")
                                continue

                        # กลับหน้าหลัก
                        safe_navigate_back(driver)

                    except Exception as e:
                        logger.error(f"Error processing suggestion {i}: {e}")
                        continue

                processed_keywords.add(keyword)
                
                # บันทึก processed keywords ทุก 5 keywords
                if len(processed_keywords) % 5 == 0:
                    save_processed_keywords(processed_keywords)

            except Exception as e:
                logger.error(f"Keyword error for {keyword}: {e}")
                continue

            # แสดงสถิติและบันทึก backup
            print_summary_stats(all_drugs, keyword, idx + 1)
            if len(all_drugs) > 0 and len(all_drugs) % 20 == 0:
                save_all_formats(all_drugs, "drug_backup.csv")
                save_processed_keywords(processed_keywords)
                logger.info(f"📝 บันทึกข้อมูล backup: {len(all_drugs)} รายการ")

            time.sleep(random.uniform(1, 3))

        # บันทึกข้อมูลสุดท้าย
        save_all_formats(all_drugs, "drug_full_details.csv")
        save_all_formats(all_drugs, "drug_backup.csv")
        save_processed_keywords(processed_keywords)
        logger.info(f"✅ บันทึกข้อมูลทั้งหมด {len(all_drugs)} รายการเรียบร้อยแล้ว")

    except KeyboardInterrupt:
        logger.info("\n⚠️ ผู้ใช้หยุดการทำงาน...")
        if 'all_drugs' in locals() and all_drugs:
            save_to_csv(all_drugs, "drug_interrupted_backup.csv")
        if 'processed_keywords' in locals():
            save_processed_keywords(processed_keywords)
    
    except Exception as e:
        logger.error(f"Critical error: {e}")
        if 'all_drugs' in locals() and all_drugs:
            save_to_csv(all_drugs, "drug_emergency_backup.csv")
        if 'processed_keywords' in locals():
            save_processed_keywords(processed_keywords)
    
    finally:
        if driver:
            try:
                driver.quit()
                logger.info("🔚 ปิด browser เรียบร้อยแล้ว")
            except:
                pass

def test_mode(mode, *args):
    """ฟังก์ชันทดสอบรวม"""
    options = setup_driver()
    driver = None
    
    try:
        driver = webdriver.Chrome(options=options)
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        
        if mode == "url" and args:
            # ทดสอบ URL โดยตรง
            logger.info(f"🧪 ทดสอบ URL: {args[0]}")
            driver.get(args[0])
            wait_for_page_load(driver)
            
            data = extract_drug_details_fixed(driver)
            print_drug_data(data)
            
            # แสดงสถิติ
            total_fields = len([k for k in data.keys() if k != "URL"])
            filled_fields = len([k for k, v in data.items() if k != "URL" and v.strip()])
            print(f"\n📊 ความสำเร็จ: {filled_fields}/{total_fields} ฟิลด์ ({(filled_fields/total_fields)*100:.1f}%)")
            
            return data
            
        elif mode == "quick":
            # ทดสอบเร็ว
            logger.info("⚡ ทดสอบเร็ว")
            driver.get("https://www.yaandyou.net/")
            wait_for_page_load(driver)

            # ข้าม SSL warning
            try:
                driver.find_element(By.ID, "details-button").click()
                time.sleep(0.5)
                driver.find_element(By.ID, "proceed-link").click()
                time.sleep(2)
            except NoSuchElementException:
                pass

            # ค้นหา keyword
            keyword = args[0] if args else "ab"
            input_box = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.ID, "keyword")))
            input_box.clear()
            input_box.send_keys(keyword)
            time.sleep(2)

            # คลิก suggestion และยาตัวแรก
            try:
                WebDriverWait(driver, 8).until(EC.presence_of_element_located((By.XPATH, '//*[@id="results"]/div')))
                suggestions = driver.find_elements(By.XPATH, '//*[@id="results"]/div')
                
                if suggestions:
                    safe_click_element(driver, suggestions[0])
                    wait_for_page_load(driver)
                    
                    for selector in ['#tableId3 tbody tr', 'table tbody tr']:
                        try:
                            rows = driver.find_elements(By.CSS_SELECTOR, selector)
                            if rows:
                                safe_click_element(driver, rows[0])
                                wait_for_page_load(driver)
                                
                                # เปิด debug ถ้าต้องการ
                                if DEBUG_MODE:
                                    debug_print("=== Quick Test Debug ===")
                                    page_text = driver.find_element(By.TAG_NAME, "body").text
                                    lines = [line.strip() for line in page_text.split('\n') if line.strip()]
                                    usage_line = next((i for i, line in enumerate(lines) if "ยานี้ใช้สำหรับ" in line), None)
                                    if usage_line:
                                        debug_print(f"Found 'ยานี้ใช้สำหรับ' at line {usage_line + 1}")
                                
                                data = extract_drug_details_fixed(driver)
                                print_drug_data(data)
                                
                                # เฉพาะการแสดงผล 'ยานี้ใช้สำหรับ'
                                usage_data = data.get("ยานี้ใช้สำหรับ", "")
                                if usage_data:
                                    bullets = usage_data.split('\n')
                                    print(f"\n🎯 'ยานี้ใช้สำหรับ': {len(bullets)} bullet points")
                                    if len(bullets) == 3:
                                        print("✅ จำนวน bullet points ถูกต้อง")
                                    else:
                                        print("❌ จำนวน bullet points ไม่ถูกต้อง (ต้องการ 3)")
                                else:
                                    print("❌ ไม่พบข้อมูล 'ยานี้ใช้สำหรับ'")
                                
                                return data
                            break
                        except:
                            continue
                            
            except TimeoutException:
                print(f"❌ ไม่พบ autocomplete สำหรับ {keyword}")

    except Exception as e:
        logger.error(f"Test error: {e}")
    finally:
        if driver:
            driver.quit()
    
    return None

def extract_usage_bullets_comprehensive(text):
    """ดึง bullet points ของการใช้ยาแบบครอบคลุม"""
    try:
        bullets = []
        seen = set()
        
        # 1. ดึงจาก HTML bullet points
        html_bullets = re.findall(r'<li>(.*?)</li>', text, re.DOTALL)
        if html_bullets:
            for bullet in html_bullets:
                clean_bullet = clean_bullet_text(bullet)
                if clean_bullet and clean_bullet not in seen:
                    bullets.append(clean_bullet)
                    seen.add(clean_bullet)
        
        # 2. ดึงจากเครื่องหมาย bullet
        if not bullets:
            bullet_markers = ['•', '·', '*', '-', '○', '●', '■', '▪', '◦']
            for marker in bullet_markers:
                if marker in text:
                    marker_bullets = [b.strip() for b in text.split(marker) if b.strip()]
                    for bullet in marker_bullets:
                        clean_bullet = clean_bullet_text(bullet)
                        if clean_bullet and clean_bullet not in seen:
                            bullets.append(clean_bullet)
                            seen.add(clean_bullet)
                    if bullets:
                        break
        
        # 3. ดึงจากประโยคที่เกี่ยวข้อง
        if not bullets:
            sentences = re.split(r'[.!?]', text)
            usage_keywords = [
                "ยานี้ใช้", "อาจใช้", "ใช้เพื่อ", "เพื่อรักษา", "รักษาโรค",
                "บรรเทาอาการ", "ช่วยควบคุม", "ช่วยลดอาการ", "ป้องกัน"
            ]
            
            for sentence in sentences:
                sentence = sentence.strip()
                if (len(sentence) > 20 and 
                    any(keyword in sentence for keyword in usage_keywords)):
                    clean_sentence = clean_bullet_text(sentence)
                    if clean_sentence and clean_sentence not in seen:
                        bullets.append(clean_sentence)
                        seen.add(clean_sentence)
        
        return bullets
        
    except Exception as e:
        debug_print(f"Bullet extraction failed: {e}")
        return []

def clean_bullet_text(text):
    """ทำความสะอาดข้อความ bullet"""
    try:
        # ลบ HTML tags
        text = re.sub(r'<[^>]+>', '', text)
        
        # ลบเครื่องหมายและคำที่ไม่ต้องการ
        text = re.sub(r'^[\*•\-\s]+', '', text)
        unwanted = ["ได้แก่", "คือ", "ประกอบด้วย", "แสดง เพิ่มเติม", "ซ่อน", "เช่น"]
        for word in unwanted:
            text = text.replace(word, "")
        
        # ทำความสะอาดช่องว่าง
        text = " ".join(text.split()).strip()
        text = text.rstrip('.,;:')
        
        return text if len(text) > 15 else ""
        
    except Exception as e:
        debug_print(f"Clean bullet failed: {e}")
        return text

def clean_single_line_content(content):
    """ทำความสะอาดเนื้อหาบรรทัดเดียว"""
    if not content:
        return ""
    
    # ลบข้อความที่ไม่ต้องการ
    unwanted = ["ซ่อน รายชื่อ", "แสดง รายชื่อ", "[-]", "[+]", "แสดง เพิ่มเติม", "ซ่อน", "(ชื่อบริษัท)"]
    for phrase in unwanted:
        content = content.replace(phrase, "")
    
    # ตัดที่หัวข้อถัดไป
    stop_indicators = ["ยานี้ใช้สำหรับ", "วิธีการใช้ยา", "สิ่งที่ควรแจ้ง"]
    for indicator in stop_indicators:
        if indicator in content:
            content = content.split(indicator)[0].strip()
            break
    
    # ทำความสะอาดช่องว่าง
    return " ".join(content.split()).strip()

def format_as_bullets(content_lines):
    """จัดรูปแบบเป็น bullet points"""
    if isinstance(content_lines, str):
        content_lines = [content_lines]
    return "\n".join([f"* {line.strip()}" for line in content_lines if line.strip()])

def extract_sections_with_patterns(lines, data):
    """ดึงข้อมูลยาแต่ละหัวข้อจาก div.bs-callout"""
    try:
        section_patterns = {
            "ชื่อสามัญ": [r"ชื่อสามัญ\s*$", r"GENERIC NAME\s*$"],
            "ชื่อการค้า": [r"ชื่อการค้า\s*$", r"TRADE NAME\s*$"],
            "รูปแบบยา": [r"รูปแบบยา\s*$", r"DOSAGE FORM\s*$", r"(?:ยา)?เจล\s*$", r"(?:ชนิด)?เจล\s*$"],
            "ยานี้ใช้สำหรับ": [r"ยานี้ใช้สำหรับ\s*$"],
            "วิธีการใช้ยา": [r"วิธีการใช้ยา\s*$"],
            "สิ่งที่ควรแจ้งให้แพทย์หรือเภสัชกรทราบ": [r"สิ่งที่ควรแจ้งให้แพทย์หรือเภสัชกรทราบ\s*$"],
            "ทำอย่างไรหากลืมรับประทานยาหรือใช้ยา": [r"ทำอย่างไรหากลืมรับประทานยาหรือใช้ยา\s*$"],
            "อาการไม่พึงประสงค์ทั่วไป": [r"อาการไม่พึงประสงค์ทั่วไป\s*$"],
            "อาการไม่พึงประสงค์ที่ต้องแจ้งแพทย์หรือเภสัชกรทันที": [r"อาการไม่พึงประสงค์ที่ต้องแจ้งแพทย์หรือเภสัชกรทันที\s*$"],
            "การเก็บรักษายา": [r"การเก็บรักษายา\s*$"]
        }
        
        # หาตำแหน่งของแต่ละหัวข้อ
        section_indices = {}
        current_section = None
        section_content = []
        
        for i, line in enumerate(lines):
            # ตรวจสอบว่าเป็นหัวข้อใหม่หรือไม่
            is_new_section = False
            new_section = None
            
            for section_name, patterns in section_patterns.items():
                for pattern in patterns:
                    if re.search(pattern, line, re.IGNORECASE):
                        is_new_section = True
                        new_section = section_name
                        break
                if is_new_section:
                    break
            
            # ถ้าเจอหัวข้อใหม่
            if is_new_section:
                # บันทึกเนื้อหาของหัวข้อก่อนหน้า
                if current_section and section_content:
                    content = clean_and_format_content(section_content)
                    if content:
                        data[current_section] = content
                
                # เริ่มเก็บเนื้อหาใหม่
                current_section = new_section
                section_content = []
            # ถ้าไม่ใช่หัวข้อและมีหัวข้อปัจจุบัน
            elif current_section and line.strip():
                # ข้ามบรรทัดที่เป็นลิงก์หรือเมนู
                if not any(skip in line for skip in ["หน้าหลัก", "ติดต่อเรา", "เงื่อนไขการใช้", "ค้นหายา"]):
                    section_content.append(line)
        
        # บันทึกเนื้อหาของหัวข้อสุดท้าย
        if current_section and section_content:
            content = clean_and_format_content(section_content)
            if content:
                data[current_section] = content
                
    except Exception as e:
        if DEBUG_MODE:
            debug_print(f"Error in extract_sections_with_patterns: {e}")

def extract_administration_bullets_comprehensive(text):
    """ดึง bullet points ของวิธีการใช้ยาแบบครอบคลุม"""
    try:
        bullets = []
        seen = set()
        
        # 1. ดึงจาก HTML bullet points
        html_bullets = re.findall(r'<li>(.*?)</li>', text, re.DOTALL)
        if html_bullets:
            for bullet in html_bullets:
                clean_bullet = clean_bullet_text(bullet)
                if clean_bullet and clean_bullet not in seen:
                    bullets.append(clean_bullet)
                    seen.add(clean_bullet)
        
        # 2. ดึงจากเครื่องหมาย bullet
        if not bullets:
            bullet_markers = ['•', '·', '*', '-', '○', '●', '■', '▪', '◦']
            for marker in bullet_markers:
                if marker in text:
                    marker_bullets = [b.strip() for b in text.split(marker) if b.strip()]
                    for bullet in marker_bullets:
                        clean_bullet = clean_bullet_text(bullet)
                        if clean_bullet and clean_bullet not in seen:
                            bullets.append(clean_bullet)
                            seen.add(clean_bullet)
                    if bullets:
                        break
        
        # 3. ดึงจากประโยคที่เกี่ยวข้อง
        if not bullets:
            sentences = re.split(r'[.!?]', text)
            admin_keywords = [
                "รับประทาน", "ฉีด", "ทา", "พ่น", "ป้าย",
                "ขนาดยา", "วิธีใช้", "ใช้ยา", "ก่อนอาหาร", "หลังอาหาร"
            ]
            
            for sentence in sentences:
                sentence = sentence.strip()
                if (len(sentence) > 20 and 
                    any(keyword in sentence for keyword in admin_keywords)):
                    clean_sentence = clean_bullet_text(sentence)
                    if clean_sentence and clean_sentence not in seen:
                        bullets.append(clean_sentence)
                        seen.add(clean_sentence)
        
        return bullets
        
    except Exception as e:
        debug_print(f"Administration bullet extraction failed: {e}")
        return []

def is_unwanted_line(line):
    """ตรวจสอบบรรทัดที่ไม่ต้องการ"""
    unwanted_patterns = [
        r"^แสดง เพิ่มเติม", r"^ซ่อน", r"^\[-\]$", r"^\[\+\]$",
        r"^รายชื่อ", r"^©.*สงวนลิขสิทธิ์",
        r"^ทรัพย์สินทางปัญญา", r"^การทำซ้ำ", r"^ข้อมูลนี้เป็นความรู้ทั่วไป"
    ]
    
    return any(re.match(pattern, line, re.IGNORECASE) for pattern in unwanted_patterns)

def clean_trade_names_in_files():
    """ทำความสะอาดชื่อการค้าในไฟล์ข้อมูลที่มีอยู่"""
    try:
        # ไฟล์ที่ต้องการทำความสะอาด
        files_to_clean = [
            "drug_full_details.json",
            "drug_backup.json",
            "drug_full_details.csv",
            "drug_backup.csv"
        ]
        
        for filename in files_to_clean:
            if not os.path.exists(filename):
                continue
                
            logger.info(f"กำลังทำความสะอาดไฟล์ {filename}")
            
            # อ่านข้อมูล
            if filename.endswith('.json'):
                with open(filename, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    
                # ทำความสะอาดชื่อการค้า
                for item in data:
                    if "ชื่อการค้า" in item:
                        item["ชื่อการค้า"] = item["ชื่อการค้า"].replace("(ชื่อบริษัท)", "").strip()
                
                # บันทึกกลับ
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                    
            elif filename.endswith('.csv'):
                rows = []
                with open(filename, 'r', encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    fieldnames = reader.fieldnames
                    for row in reader:
                        if "ชื่อการค้า" in row:
                            row["ชื่อการค้า"] = row["ชื่อการค้า"].replace("(ชื่อบริษัท)", "").strip()
                        rows.append(row)
                
                # บันทึกกลับ
                with open(filename, 'w', encoding='utf-8-sig', newline='') as f:
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(rows)
            
            logger.info(f"ทำความสะอาดไฟล์ {filename} เสร็จสิ้น")
        
        # สร้างไฟล์ txt ใหม่
        if os.path.exists("drug_full_details.json"):
            with open("drug_full_details.json", 'r', encoding='utf-8') as f:
                data = json.load(f)
            save_to_txt(data, "drug_full_details.txt")
            
        if os.path.exists("drug_backup.json"):
            with open("drug_backup.json", 'r', encoding='utf-8') as f:
                data = json.load(f)
            save_to_txt(data, "drug_backup.txt")
            
        logger.info("ทำความสะอาดไฟล์ทั้งหมดเสร็จสิ้น")
        return True
        
    except Exception as e:
        logger.error(f"เกิดข้อผิดพลาดในการทำความสะอาดไฟล์: {e}")
        return False

def clean_existing_data_content():
    """ทำความสะอาดเนื้อหาในไฟล์ข้อมูลที่มีอยู่"""
    try:
        # ไฟล์ที่ต้องการทำความสะอาด
        files_to_clean = [
            "drug_full_details.json",
            "drug_backup.json",
            "drug_full_details.csv",
            "drug_backup.csv"
        ]
        
        # หัวข้อที่ต้องการทำความสะอาด
        sections_to_clean = [
            "อาการไม่พึงประสงค์ทั่วไป",
            "อาการไม่พึงประสงค์ที่ต้องแจ้งแพทย์หรือเภสัชกรทันที"
        ]
        
        # รูปแบบที่ต้องการลบ (เรียงจากยาวไปสั้น)
        patterns_to_remove = [
            r"^\*?\s*(?:มีอาการไม่พึงประสงค์ทั่วไป|อาการไม่พึงประสงค์ทั่วไป)?\s*(?:มี)?(?:ดังนี้|ต่อไปนี้|ดังต่อไปนี้)\s*[:：]*",
            r"^\*?\s*(?:มีอาการ|อาการ)?\s*(?:มี)?(?:ดังนี้|ต่อไปนี้|ดังต่อไปนี้)\s*[:：]*",
            r"^\*?\s*(?:มี)?(?:ดังนี้|ต่อไปนี้|ดังต่อไปนี้)\s*[:：]*",
            r"^\*?\s*(?:ได้แก่|เช่น|อาทิ)\s*[:：]*",
            r"^\*?\s*(?:พบได้|อาจพบ|มักพบ)\s*[:：]*",
            r"^[\*\s]*$",  # บรรทัดที่มีแต่ * และช่องว่าง
            r"^\*\s*$",    # บรรทัดที่มีแค่ *
            r"^\s*\*\s*$", # บรรทัดที่มีแค่ * และอาจมีช่องว่าง
            r"^มีดังนี้\s*[:：]*",  # เพิ่มรูปแบบใหม่
            r"^มี\s*ดังนี้\s*[:：]*",  # เพิ่มรูปแบบใหม่
            r"^ดังนี้\s*[:：]*",  # เพิ่มรูปแบบใหม่
            r"^\*\s*มีดังนี้\s*[:：]*",  # เพิ่มรูปแบบใหม่
            r"^\*\s*มี\s*ดังนี้\s*[:：]*",  # เพิ่มรูปแบบใหม่
            r"^\*\s*ดังนี้\s*[:：]*"  # เพิ่มรูปแบบใหม่
            r"^\*\sมีดังนี้\s*[:：]*"  # เพิ่มรูปแบบใหม่
        ]
        
        def clean_line_content(line):
            """ทำความสะอาดเนื้อหาแต่ละบรรทัด"""
            if not line or line.isspace():
                return ""
                
            # ลบคำที่ไม่ต้องการออก
            clean_line = line.strip()
            for pattern in patterns_to_remove:
                clean_line = re.sub(pattern, "", clean_line, flags=re.IGNORECASE).strip()
            
            # ถ้าเหลือแค่เครื่องหมายวรรคตอนหรือช่องว่าง ให้ข้าม
            if not clean_line or clean_line in [":", "：", "-", ".", ",", ";"]:
                return ""
            
            # ลบ * ที่อาจเหลืออยู่ที่ต้นบรรทัด
            clean_line = re.sub(r'^\*\s*', '', clean_line).strip()
            
            # ถ้ายังมีเนื้อหาเหลืออยู่ ให้เพิ่ม * นำหน้า
            return f"* {clean_line}" if clean_line else ""
        
        for filename in files_to_clean:
            if not os.path.exists(filename):
                continue
                
            logger.info(f"กำลังทำความสะอาดไฟล์ {filename}")
            
            # อ่านข้อมูล
            if filename.endswith('.json'):
                with open(filename, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    
                # ทำความสะอาดเนื้อหา
                for item in data:
                    for section in sections_to_clean:
                        if section in item and isinstance(item[section], str):
                            lines = item[section].split('\n')
                            cleaned_lines = []
                            
                            for line in lines:
                                clean_line = clean_line_content(line)
                                if clean_line:
                                    cleaned_lines.append(clean_line)
                            
                            item[section] = '\n'.join(cleaned_lines) if cleaned_lines else ""
                
                # บันทึกกลับ
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                    
            elif filename.endswith('.csv'):
                rows = []
                with open(filename, 'r', encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    fieldnames = reader.fieldnames
                    for row in reader:
                        for section in sections_to_clean:
                            if section in row and isinstance(row[section], str):
                                lines = row[section].split('\n')
                                cleaned_lines = []
                                
                                for line in lines:
                                    clean_line = clean_line_content(line)
                                    if clean_line:
                                        cleaned_lines.append(clean_line)
                                
                                row[section] = '\n'.join(cleaned_lines) if cleaned_lines else ""
                        rows.append(row)
                
                # บันทึกกลับ
                with open(filename, 'w', encoding='utf-8-sig', newline='') as f:
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(rows)
            
            logger.info(f"ทำความสะอาดไฟล์ {filename} เสร็จสิ้น")
        
        # สร้างไฟล์ txt ใหม่
        if os.path.exists("drug_full_details.json"):
            with open("drug_full_details.json", 'r', encoding='utf-8') as f:
                data = json.load(f)
            save_to_txt(data, "drug_full_details.txt")
            
        if os.path.exists("drug_backup.json"):
            with open("drug_backup.json", 'r', encoding='utf-8') as f:
                data = json.load(f)
            save_to_txt(data, "drug_backup.txt")
            
        logger.info("ทำความสะอาดไฟล์ทั้งหมดเสร็จสิ้น")
        return True
        
    except Exception as e:
        logger.error(f"เกิดข้อผิดพลาดในการทำความสะอาดไฟล์: {e}")
        return False

# เพิ่มตัวเลือกในฟังก์ชัน main
if __name__ == "__main__":
    if len(sys.argv) > 1:
        command = sys.argv[1]
        
        if command == "clean":
            # ทำความสะอาดเนื้อหาในไฟล์ที่มีอยู่
            clean_existing_data_content()
            
        elif command == "debug":
            set_debug_mode(True)
            if len(sys.argv) > 2:
                if sys.argv[2] == "url" and len(sys.argv) > 3:
                    test_mode("url", sys.argv[3])
                else:
                    test_mode("quick", sys.argv[2] if len(sys.argv) > 2 else "aa")
            else:
                main()
                
        elif command == "quick":
            keyword = sys.argv[2] if len(sys.argv) > 2 else "aa"
            test_mode("quick", keyword)
            
        elif command == "url" and len(sys.argv) > 2:
            test_mode("url", sys.argv[2])
            
        elif command == "usage":
            # ทดสอบการแยกข้อมูลด้วยข้อความตัวอย่าง
            test_text = """
            ยานี้ใช้เพื่อรักษาโรคติดเชื้อแบคทีเรีย เช่น โรคปอดอักเสบ โรคหลอดลมอักเสบ โรคติดเชื้อที่หู จมูก ลำคอ ผิวหนัง ระบบทางเดินปัสสาวะ โรคหนองในเทียมคลามัยเดีย ทราโคมาทิส (chlamydia trachomatis) (เป็นยาทางเลือกในหญิงตั้งครรภ์ที่แพทย์พิจารณาถึงความจำเป็นแล้วเท่านั้น)
            อาจใช้ร่วมกับยาอื่นเพื่อรักษาโรคกระเพาะอาหารจากการติดเชื้อเฮลิโคแบคเตอร์ไพโลไร (H.pylori)
            ยานี้อาจใช้เพื่อรักษาโรคหรืออาการอื่นๆได้ หากมีข้อสงสัยควรปรึกษาแพทย์หรือเภสัชกร
            """
            
            print("🧪 ทดสอบการแยกข้อมูล 'ยานี้ใช้สำหรับ'")
            set_debug_mode(True)
            
            bullets = extract_usage_bullets_comprehensive(test_text.strip())
            result = "\n".join([f"* {bullet}" for bullet in bullets])
            
            print("\n📋 ผลลัพธ์:")
            print(result)
            print(f"\n📊 จำนวน bullet points: {len(bullets)}")
            
        elif command == "help":
            print("📖 วิธีใช้งาน:")
            print("  python script.py                    # เริ่มการทำงานปกติ")
            print("  python script.py quick [keyword]    # ทดสอบเร็ว (default: aa)")
            print("  python script.py url [URL]          # ทดสอบ URL โดยตรง")
            print("  python script.py debug [keyword]    # โหมด debug")
            print("  python script.py debug url [URL]    # debug URL เฉพาะ")
            print("  python script.py usage              # ทดสอบการแยก 'ยานี้ใช้สำหรับ'")
            print("  python script.py help               # แสดงความช่วยเหลือ")
        else:
            main()
    else:
        main()